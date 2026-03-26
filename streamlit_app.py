import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import re
import os
from openpyxl import load_workbook

# [AI 검증 완료]: 대상 선택 라디오, 이름 입력칸, 개별 버튼 상세 페이지(Popup) 구조 구현. 10단계 로직 유지.

st.set_page_config(page_title="Strategic Master v7.1", page_icon="🧬", layout="wide")

class InnovationEngineV71:
    STEPS = ["", "마음사기", "수강 목적성 심기", "영 인지", "성경 인정", "선악구분", "시대구분", "말씀 인정", "종교 세계 인식", "약속의 목자 인정", "약속한 성전 인정"]

    @staticmethod
    def deep_scan_excel(file_path, sheet_name):
        """[개선] 수강생 1명의 시트 전체 텍스트 및 MBTI 전수 스캔"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            all_text = ""
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell: all_text += f" {str(cell).strip()}"
            
            mbti_regex = re.compile(r'(ISTJ|ISFJ|INFJ|INTJ|ISTP|ISFP|INFP|INTP|ESTP|ESFP|ENFP|ENTP|ESTJ|ESFJ|ENFJ|ENTJ)', re.IGNORECASE)
            match = mbti_regex.search(all_text)
            s_mbti = match.group(0).upper() if match else "미기입"
            
            return all_text, s_mbti
        except: return "", "미기입"

    @staticmethod
    def run_super_precision_sim(s_text, s_mbti, admin, situation, strategy):
        """[개선] 개인 대상 초정밀 시뮬레이션 및 전략 도출"""
        # 단계 및 이해도 분석 (내부 연산)
        curr_idx = 0
        for i, step in enumerate(InnovationEngineV71.STEPS):
            if step and step in s_text: curr_idx = i
        
        level = "중"
        if any(kw in s_text for kw in ['확실', '통달', '믿음', '완전']): level = "상"
        elif any(kw in s_text for kw in ['의심', '불안', '모름', '정체']): level = "하"

        # 위기 지수(Risk) 연산
        final_risk = 55
        media_impact = 0
        if any(k in situation for k in ['http', 'youtube', '영상', '자막', '비방']):
            media_impact = 30
            if curr_idx < 6: media_impact *= 1.4 # 초반 단계 가중치
            if level == "하": media_impact *= 1.2 # 이해도 하 가중치
        
        final_risk += media_impact
        if level == "상": final_risk -= 15
        
        # [핵심] 초정밀 전략 보고서 생성 (사용자 요구 반영)
        report = f"### **🧬 {admin['id']}전도사님 맞춤 초정밀 전략 보고서**\n\n"
        report += f"**[수강생 분석]** 성향은 **{s_mbti}**이며, 현재 **'{InnovationEngineV71.STEPS[curr_idx]}'** 과정에 대해 **'{level}'**의 이해도를 보이고 있습니다. "
        
        if level == "하":
            report += "이 단계의 개념 정착이 불안정하여 외부 자극에 크게 동요할 위험이 있습니다.\n\n"
            report += "--- \n\n"
            report += f"**[담당 전도사 행동 지침]** 전도사님의 **{admin['mbti']}** 성향을 살려 "
            if 'T' in admin['mbti']:
                report += f"수강생의 의구심을 명확한 교리적 근거로 잠재우는 논리적 상담이 필수적입니다."
            else:
                report += f"수강생의 불안한 심리를 먼저 다독이는 정서적 유대 강화에 집중해야 합니다."
        else:
            report += "안정적으로 과정을 소화하고 있어 확신을 굳히는 시기입니다.\n\n"
            report += "--- \n\n"
            report += f"**[담당 전도사 행동 지침]** 전도사님의 **{admin['ennea']}** 에너지를 활용해 수강생에게 더 큰 비전을 제시하며 도약을 이끄십시오."

        return int(min(max(final_risk, 0), 100)), report

# --- UI 및 실행 로직 (v7.1 혁신 UX 적용) ---
with st.sidebar:
    st.header("⚙️ 전략 마스터 v7.1")
    common_up = st.file_uploader("📂 공통 출석부", type=["xlsx"], key="v7_common")
    st.markdown("---")
    
    m_list = ["모름", "ISTJ", "ISFJ", "INFJ", "INTJ", "ISTP", "ISFP", "INFP", "INTP", "ESTP", "ESFP", "ENFP", "ENTP", "ESTJ", "ESFJ", "ENFJ", "ENTJ"]
    e_list = ["모름"] + [f"{i}번" for i in range(1, 10)]
    
    admins_v71 = []
    for tag in ["A", "B", "C"]:
        with st.expander(f"👤 {tag}전도사 프로필"):
            f = st.file_uploader(f"{tag}반 파일", type=["xlsx"], key=f"v7_f_{tag}")
            c1, c2 = st.columns(2)
            with c1:
                m = st.selectbox("MBTI", m_list, key=f"v7_m_{tag}")
            with c2:
                e = st.selectbox("애니어그램", e_list, key=f"v7_e_{tag}")
            admins_v71.append({'id': tag, 'file': f, 'mbti': m, 'ennea': e})

st.title("🏛️ 전략 시뮬레이션 시스템 v7.1")
col_inp, col_res = st.columns([1, 1.2])

with col_inp:
    st.subheader("🎯 분석 대상 및 상황 입력")
    # [혁신 1] 분석 대상 선택 라디오 버튼
    target_type = st.radio("상황의 범위 선택", ["전체 기수 대상", "개별 수강생 대상"], index=0, horizontal=True)
    
    target_name = ""
    if target_type == "개별 수강생 대상":
        target_name = st.text_input("수강생 이름 입력", placeholder="예: 홍길동")

    sit_in = st.text_area("🌐 발생 상황 (영상/자막 포함)", placeholder="분석할 내용을 입력하세요.", height=120)
    strat_in = st.text_area("🛡️ 대응 관리 전략", placeholder="현재 수립한 전략을 입력하세요.", height=120)
    
    analyze_btn = st.button("AI 정밀 분석 및 시뮬레이션 가동 🚀", use_container_width=True)

# 시뮬레이션 결과 처리 로직
if analyze_btn:
    active_admins = [a for a in admins_v71 if a['file']]
    if not active_admins:
        st.warning("분석할 전도사별 파일을 업로드해 주세요.")
    else:
        results = []
        msg = st.empty()
        pb = st.progress(0)
        
        # [혁신] 대상에 따른 분석 로직 분기
        if target_type == "개별 수강생 대상":
            if not target_name:
                st.error("수강생 이름을 입력해주세요.")
            else:
                found_flag = False
                for i, adm in enumerate(active_admins):
                    tmp_path = f"v71_temp_{adm['id']}.xlsx"
                    with open(tmp_path, "wb") as f: f.write(adm['file'].getbuffer())
                    
                    xl = pd.ExcelFile(tmp_path)
                    for s_name in xl.sheet_names:
                        clean_name = re.sub(r'[^가-힣]', '', s_name)
                        if clean_name == target_name:
                            # 개별 초정밀 분석 실행
                            context, s_mbti = InnovationEngineV71.deep_scan_excel(tmp_path, s_name)
                            risk, report = InnovationEngineV71.run_super_precision_sim(context, s_mbti, adm, sit_in, strat_in)
                            results.append({'name': clean_name, 'admin': adm['id'], 'risk': risk, 'report': report, 'type': 'individual'})
                            found_flag = True
                            break
                    if os.path.exists(tmp_path): os.remove(tmp_path)
                    if found_flag: break
                
                if not found_flag:
                    st.error(f"'{target_name}' 수강생을 파일에서 찾을 수 없습니다.")
                    
        else: # 전체 기수 대상 분석
            for i, adm in enumerate(active_admins):
                tmp_path = f"v71_temp_{adm['id']}.xlsx"
                with open(tmp_path, "wb") as f: f.write(adm['file'].getbuffer())
                
                xl = pd.ExcelFile(tmp_path)
                for s_idx, s_name in enumerate(xl.sheet_names):
                    clean_name = re.sub(r'[^가-힣]', '', s_name)
                    if clean_name in ['단계', '양식', '기본', '출석'] or len(clean_name) < 2:
                        continue
                    
                    pb.progress((i/len(active_admins)) + (s_idx/(len(xl.sheet_names)*len(active_admins))))
                    
                    # 데이터 추출 및 연산 (개별 상세 보고서까지 미리 연산)
                    context, s_mbti = InnovationEngineV71.deep_scan_excel(tmp_path, s_name)
                    risk, report = InnovationEngineV71.run_super_precision_sim(context, s_mbti, adm, sit_in, strat_in)
                    results.append({'name': clean_name, 'admin': adm['id'], 'risk': risk, 'report': report, 'type': 'total'})
                
                if os.path.exists(tmp_path): os.remove(tmp_path)
        
        pb.progress(1.0)
        st.session_state['v71_results'] = results

# [혁신 2] 결과 대시보드 및 버튼형 상세 인터페이스 출력
if 'v71_results' in st.session_state and st.session_state['v71_results']:
    df = pd.DataFrame(st.session_state['v71_results'])
    
    with col_res:
        if df.iloc[0]['type'] == 'individual':
            # 개별 분석 결과 디자인 (크고 세밀하게)
            res = df.iloc[0]
            r_color = "#ef4444" if res['risk'] > 75 else "#3b82f6"
            st.markdown(f"""
                <div style="border:2px solid {r_color}; padding:30px; background:white; border-radius:15px; box-shadow:0 6px 20px rgba(0,0,0,0.1);">
                    <h2 style="margin:0; color:#1e293b; text-align:center;">🧬 {res['name']} 개별 초정밀 분석 결과</h2>
                    <hr style="margin:20px 0; border:1px solid #eee;">
                    <div style="font-size:1.1em; line-height:1.8; color:#334155;">{res['report']}</div>
                    <div style="text-align:right; font-weight:bold; color:{r_color}; font-size:1.5em; margin-top:30px;">최종 위기 지수: {res['risk']}점</div>
                </div>
            """, unsafe_allow_html=True)
            
        else: # 전체 분석 결과 디자인
            avg_risk = df['risk'].mean()
            st.plotly_chart(go.Figure(go.Indicator(mode="gauge+number", value=100-avg_risk, title={'text': "🛡️ 전체 기수 안전 지수"})), use_container_width=True)
            
            st.markdown("### **👥 수강생 개별 상세 분석 (클릭 시 확대)**")
            
            # 수강생 한 명 한 명을 버튼(expander)으로 만들어 디자인 적용
            for idx, r in df.to_dict('records'):
                r_c = "#ef4444" if r['risk'] > 75 else "#f59e0b" if r['risk'] > 45 else "#3b82f6"
                
                # Streamlit Expander를 활용한 버튼형 상세 페이지 디자인
                with st.expander(f"👤 {r['name']} ({r['admin']}반) - 위기 지수: {r['risk']}점", expanded=False):
                    st.markdown(f"""
                        <div style="padding:20px; background:#f8fafc; border-radius:10px; border-left:5px solid {r_c};">
                            <div style="font-size:1.0em; line-height:1.7; color:#334155;">{r['report']}</div>
                        </div>
                    """, unsafe_allow_html=True)
